import serial
import  sys
import openpyxl
import threading
# initialize excel
wb = openpyxl.Workbook()
ws = wb.active
ws.cell( row = 1,column =1,value = "s_rate1")
ws.cell( row = 1,column =2,value = "rate1")
ws.cell( row = 1,column =3,value = "s_rate2")
ws.cell( row = 1,column =4,value = "rate2")
i= 1
s_rate1 =0
s_rate2 =0
s_pos1 =0
s_pos2 =0
mess = ""
# initialize UART
ser = serial.Serial( port="/dev/ttyUSB0", baudrate=115200)



def processData(data):
    data = data.replace("!", "")
    data = data.replace("#", "")
    splitData = data.split(";")
    print(splitData)
    try:
        if splitData[0] == 'V':
            global i,s_rate1,s_rate2
            i = i +1
            v1 = int(splitData[1])
            v2 = int(splitData[2])
            #print('V1 =',v1,'\n')
            #print ('V2 =',v2,'\n')
            ws.cell(row = i, column = 2,value= v1)
            ws.cell(row = i, column = 4,value= v2)
            ws.cell(row = i, column = 1, value =  s_rate1)
            ws.cell(row = i, column = 3, value =  s_rate2)
        elif splitData[0] =='P':
            p1 = int(splitData[1])
            p2 = int(splitData[2])
            #print('p1 =',p1,'\n')
            #print ('p2 =',p2,'\n')
    except:
        pass

def readSerial():
    bytesToRead = ser.inWaiting()
    if (bytesToRead > 0):
        global mess
        try:
            mess = mess + ser.read(bytesToRead).decode("UTF-8")
            while ("#" in mess) and ("!" in mess):
                start = mess.find("!")
                end = mess.find("#")
                processData(mess[start:end + 1])
                if (end == len(mess)):
                    mess = ""
                else:
                    mess = mess[end+1:]
        except :
            print("Noise ! \n")
    #print(mess)
def input_command():
    
    try:
        while True:
            global s_rate1,s_rate2,s_pos1,s_pos2
            command =input()
            data = command
            data = data.replace("!", "")
            data = data.replace("#", "")
            ser.write((command).encode())
            if (data[0] == 'V'):
                splitData = data.split(";")
                s_rate1 = str(splitData[0])
                s_rate2 = str(splitData[1])
            elif (data[0] == 'P'):
                splitData = data.split(";")
                s_pos1 = str(splitData[0])
                s_pos2 = str(splitData[1]) 
    except KeyboardInterrupt:
         pass   
            
# main
thread = threading.Thread(target = input_command)
thread.start()

                
try:
    while True:
        readSerial()
except KeyboardInterrupt:
    wb.save("test2.xlsx")