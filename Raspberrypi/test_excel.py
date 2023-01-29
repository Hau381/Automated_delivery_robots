import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.cell( row = 1,column =1,value = "s_rate1")
ws.cell( row = 1,column =2,value = "rate1")
ws.cell( row = 1,column =3,value = "s_rate2")
ws.cell( row = 1,column =4,value = "rate2")


wb.save("test.xlsx")